'''
Created on Oct 28, 2013

@author: hemantbhonsle
'''
from __future__ import print_function

import xlrd
from xlrd.timemachine import REPR
import openpyxl
from openpyxl import load_workbook
import sys
import glob

book = xlrd.open_workbook("Marketplace_premium_databook_2014.xlsx")
sheet = book.sheet_by_name("Zip Code-Rating Area Lookup")
bookObama = load_workbook("Obamacare Health Insurance 1-15_updated.xlsx")

def get_zip_code_info(zip):
    for row_num in xrange(sheet.nrows):
        if sheet.row(row_num)[3].value == str(zip):
            return sheet.row(row_num)
        
def get_state_name(zip):
    for row_num in xrange(sheet.nrows):
        if sheet.row(row_num)[3].value == str(zip):
            return sheet.row(row_num)[0].value
        
def get_county_code(zip):
    for row_num in xrange(sheet.nrows):
        if sheet.row(row_num)[3].value == str(zip):
            return sheet.row(row_num)[1].value
        
def get_county_name(zip):
    for row_num in xrange(sheet.nrows):
        if sheet.row(row_num)[3].value == str(zip):
            return sheet.row(row_num)[2].value
        
def get_rating_area(zip):
    for row_num in xrange(sheet.nrows):
        if sheet.row(row_num)[3].value == str(zip):
            return sheet.row(row_num)[4].value

# stateInitials must be entered in '', and rating should be entered as x.0, so 1 is 1.0
def get_state_rating_info(stateInitials, rating):
    sheet = book.sheet_by_name(stateInitials)
    for row_num in xrange(sheet.nrows):
        print(sheet.row(row_num)[3].value)
        if sheet.row(row_num)[3].value == rating:
            return sheet.row(row_num)

def safs(stateInitials, age):
    if age > 0 and age < 21:
        wb=load_workbook('0to20.xlsx')
    elif age > 20 and age < 23:
        wb=load_workbook('21.xlsx')
    elif age > 22 and age < 28:
        wb=load_workbook('25.xlsx')
    elif age > 27 and age < 33:
        wb=load_workbook('30.xlsx')
    elif age > 32 and age < 38:
        wb=load_workbook('35.xlsx')
    elif age > 37 and age < 43:
        wb=load_workbook('40.xlsx')
    elif age > 42 and age < 48:
        wb=load_workbook('45.xlsx')
    elif age > 47 and age < 53:
        wb=load_workbook('50.xlsx')
    elif age > 52 and age < 58:
        wb=load_workbook('55.xlsx')
    elif age > 57:
        wb=load_workbook('60.xlsx')
    ws=wb.get_sheet_by_name(stateInitials)
#    ws.cell('F2').value= str(age)
#    wb.save('Marketplace_premium_databook_2014 (2).xlsx')

def getSutterHealthPlan(tier, age, area):
    sheetObama = bookObama.get_sheet_by_name("Shutter Health Plan")
    base = 0
    if area > 3 and area != 10:
        return 'area not covered'
    elif area == 1:
        base = 11
    elif area == 2:
        base = 56
    elif area == 3:
        base = 101
    elif area == 10:
        base = 146
    if age > 20:
        base = base + age - 20
    base2 = str(base)
    if str(tier) == 'platinum':
        return sheetObama.cell('D' + base2).value
    elif str(tier) == 'gold':
        return sheetObama.cell('E' + base2).value
    elif str(tier) == 'silver':
        return sheetObama.cell('F' + base2).value
    elif str(tier) == 'bronze':
        return sheetObama.cell('G' + base2).value

def getVenturaHealthPlan(tier, age):
    sheetObama = bookObama.get_sheet_by_name("Ventura County Health Care Plan")
    base = age + 7
    base2 = str(base)
    if str(tier) == 'platinum':
        return sheetObama.cell('B' + base2).value
    elif str(tier) == 'gold':
        return sheetObama.cell('C' + base2).value
    elif str(tier) == 'silver':
        return sheetObama.cell('D' + base2).value
    elif str(tier) == 'bronze':
        return sheetObama.cell('E' + base2).value

#this method is breaking on call to cell attribute, no idea why.
def getValletHealthPlan(tier, age):
    sheetObama = bookObama.get_sheet_by_name("Vallet Health Plan")
    base = 21
    if age > 20:
        base = age + base - 20
    base2 = str(base)
    if str(tier) == 'platinum':
        return sheetObama.cell('CT' + base2).value
    elif str(tier) == 'gold':
        return sheetObama.cell('CU' + base2).value
    elif str(tier) == 'silver':
        return sheetObama.cell('CV' + base2).value
    elif str(tier) == 'bronze':
        return sheetObama.cell('CR' + base2).value
    elif str(tier) == 'catastrophic':
        return sheetObama.cell('CS' + base2).value

def getContraHealthPlan(tier, age):
    sheetObama = bookObama.get_sheet_by_name("Contra Costa Health Plan")
    base = 21
    if age > 20:
        base = age + base - 20
    base2 = str(base)
    if str(tier) == 'platinum':
        return sheetObama.cell('BP' + base2).value
    elif str(tier) == 'gold':
        return sheetObama.cell('BQ' + base2).value
    elif str(tier) == 'silver':
        return sheetObama.cell('BR' + base2).value
    elif str(tier) == 'bronze':
        return sheetObama.cell('BN' + base2).value
    elif str(tier) == 'catastrophic':
        return sheetObama.cell('BO' + base2).value

def getSharpHealthPlan(tier, age):
    sheetObama = bookObama.get_sheet_by_name("Sharp Health Plan")
    base = 21
    if age > 20:
        base = age + base - 20
    base2 = str(base)
    if str(tier) == 'platinum':
        return sheetObama.cell('JM' + base2).value
    elif str(tier) == 'gold':
        return sheetObama.cell('JN' + base2).value
    elif str(tier) == 'silver':
        return sheetObama.cell('JO' + base2).value
    elif str(tier) == 'bronze':
        return sheetObama.cell('JP' + base2).value
    elif str(tier) == 'catastrophic':
        return sheetObama.cell('JQ' + base2).value

#only has platinum gold and silver
def getCaNetHealthPlan(tier, age):
    sheetObama = bookObama.get_sheet_by_name("Health Net of California")
    base = 12
    if age > 20:
        base = age + base - 20
    base2 = str(base)
    if str(tier) == 'platinum':
        return sheetObama.cell('HY' + base2).value
    elif str(tier) == 'gold':
        return sheetObama.cell('HZ' + base2).value
    elif str(tier) == 'silver':
        return sheetObama.cell('IA' + base2).value

#can only take in rating areas 4 and 8
def getChineseHealthPlan(tier, age, area):
    sheetObama = bookObama.get_sheet_by_name("Chinese Community Health Plan")
    base = 0
    if area != 4 and area != 8:
        return 'area not covered'
    elif area == 4:
        base = 13
    elif area == 8:
        base = 63
    
    if age > 20:
        base = base + age - 20
    base2 = str(base)

    if str(tier) == 'platinum':
        return sheetObama.cell('B' + base2).value
    elif str(tier) == 'gold':
        return sheetObama.cell('E' + base2).value
    elif str(tier) == 'silver':
        return sheetObama.cell('H' + base2).value
    elif str(tier) == 'bronze':
        return sheetObama.cell('K' + base2).value
    elif str(tier) == 'catastrophic':
        return sheetObama.cell('N' + base2).value

#can only take in rating areas 15,16,17, and 19
def getMolinaHealthPlan(tier, age, area):
    sheetObama = bookObama.get_sheet_by_name("Molina Healthcare of California")
    base = 0
    if area != 15 and area != 16 and area != 17 and area != 19:
        return 'area not covered'
    elif area == 15:
        base = 8
    elif area == 16:
        base = 53
    elif area == 17:
        base = 98
    elif area == 19:
        base = 143

    if age > 20:
        base = base + age - 20
    base2 = str(base)
    
    if str(tier) == 'platinum':
        return sheetObama.cell('A' + base2).value
    elif str(tier) == 'gold':
        return sheetObama.cell('B' + base2).value
    elif str(tier) == 'silver':
        return sheetObama.cell('C' + base2).value
    elif str(tier) == 'bronze':
        return sheetObama.cell('D' + base2).value
    elif str(tier) == 'catastrophic':
        return sheetObama.cell('E' + base2).value






